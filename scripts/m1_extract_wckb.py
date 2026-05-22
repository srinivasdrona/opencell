"""Extract M1-relevant values from WholeCellKB data.xlsx."""

import openpyxl

XLSX = r"E:\opencell\data\m1_sources\WholeCellKB\public\fixtures\data.xlsx"
wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)

# --- Show ALL Metabolites columns with full headers (rows 1+2 merged) ---
ws = wb["Metabolites"]
rows = list(ws.iter_rows(values_only=True, max_row=2))
print("=== Metabolites: all columns ===")
for i in range(len(rows[0])):
    print(f"  [{i:2}] section={rows[0][i]!r}  name={rows[1][i]!r}")

print()
print("=== Find ATP across all columns ===")
for ri, row in enumerate(ws.iter_rows(values_only=True)):
    if ri < 2:
        continue
    if row[0] == "ATP":
        for ci, v in enumerate(row):
            if v not in (None, "", False):
                print(f"  col[{ci}] = {v!r}")
        break

# --- Misc params: print all 16 columns of header for clarity ---
print()
ws4 = wb["Misc. parameters"]
hdr_rows = list(ws4.iter_rows(values_only=True, max_row=2))
print("=== Misc. parameters: all columns ===")
for i in range(len(hdr_rows[0])):
    print(f"  [{i:2}] section={hdr_rows[0][i]!r}  name={hdr_rows[1][i]!r}")
